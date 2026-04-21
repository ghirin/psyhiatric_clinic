
import csv
import json
import pandas as pd
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import models
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy

@login_required
def patient_delete(request, pk):
    """Удаление пациента с подтверждением"""
    patient = get_object_or_404(Patient, pk=pk)
    if request.method == 'POST':
        patient.delete()
        messages.success(request, 'Пациент удалён')
        return redirect('patients:patient_list')
    return render(request, 'patients/patient_confirm_delete.html', {'patient': patient})
from django.utils import timezone
from django.views.decorators.http import require_POST, require_http_methods
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .forms import PatientForm, PatientSearchForm, PatientExportForm
from .models import Patient, Diagnosis
from users.mixins import RoleRequiredMixin, PermissionRequiredMixin, ObjectPermissionMixin, role_required, permission_required

@login_required
def patient_update(request, pk):
    """Редактирование пациента"""
    patient = get_object_or_404(Patient, pk=pk)
    if request.method == 'POST':
        form = PatientForm(request.POST, instance=patient)
        if form.is_valid():
            form.save()
            messages.success(request, 'Данные пациента обновлены')
            return redirect('patients:patient_detail', pk=patient.pk)
    else:
        form = PatientForm(instance=patient)
    return render(request, 'patients/patient_form.html', {'form': form, 'patient': patient})

@login_required
def patient_create(request):
    """Создание нового пациента"""
    if request.method == 'POST':
        form = PatientForm(request.POST)
        if form.is_valid():
            patient = form.save(commit=False)
            patient.created_by = request.user
            patient.save()
            messages.success(request, 'Пациент успешно добавлен')
            return redirect('patients:patient_detail', pk=patient.pk)
    else:
        form = PatientForm()
    return render(request, 'patients/patient_form.html', {'form': form})

@login_required
def patient_list(request):
    """Список пациентов с поиском и фильтрацией"""
    form = PatientSearchForm(request.GET or None)
    
    # Базовый queryset с учетом прав доступа
    if request.user.is_administrator or request.user.has_perm('patients.view_all_patients'):
        patients = Patient.objects.all()
    elif request.user.is_doctor:
        patients = Patient.objects.filter(attending_physician=request.user)
    else:
        # Медсестры, регистраторы, аналитики видят всех пациентов
        patients = Patient.objects.all()
    
    patients = patients.select_related('attending_physician')
    
    # Применяем фильтры
    if form.is_valid():
        query = form.cleaned_data.get('query')
        status = form.cleaned_data.get('status')
        gender = form.cleaned_data.get('gender')
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')

        if query:
            patients = patients.filter(
                Q(last_name__icontains=query) |
                Q(first_name__icontains=query) |
                Q(middle_name__icontains=query) |
                Q(case_number__icontains=query) |
                Q(passport_series__icontains=query) |
                Q(passport_number__icontains=query) |
                Q(inn__icontains=query) |
                Q(phone__icontains=query) |
                Q(address__icontains=query)
            )
        if status:
            patients = patients.filter(status=status)
        if gender:
            patients = patients.filter(gender=gender)
        if date_from:
            patients = patients.filter(admission_date__date__gte=date_from)
        if date_to:
            patients = patients.filter(admission_date__date__lte=date_to)
    # Статистика по статусам с учетом прав доступа
    status_counts = {}
    for status_code, status_name in Patient.STATUS_CHOICES:
        if request.user.is_administrator or request.user.has_perm('patients.view_all_patients'):
            status_counts[status_code] = Patient.objects.filter(status=status_code).count()
        elif request.user.is_doctor:
            status_counts[status_code] = Patient.objects.filter(
                status=status_code,
                attending_physician=request.user
            ).count()
        else:
            status_counts[status_code] = Patient.objects.filter(status=status_code).count()

    # Для каждого пациента вычисляем права
    patient_permissions = []
    for patient in patients:
        can_edit = patient.user_can_edit(request.user)
        can_delete = patient.user_can_delete(request.user)
        can_discharge = (patient.status == 'HOSPITALIZED' and can_edit)
        patient_permissions.append({
            'patient': patient,
            'can_edit': can_edit,
            'can_delete': can_delete,
            'can_discharge': can_discharge,
        })

    context = {
        'patients_with_perms': patient_permissions,
        'form': form,
        'total_patients': patients.count(),
        'sort_by': request.GET.get('sort', '-admission_date'),
        'status_counts': status_counts,
        'can_create_patient': (
            request.user.is_administrator or 
            request.user.is_doctor or 
            request.user.is_nurse or 
            request.user.is_registrar
        ),
    }
    return render(request, 'patients/patient_list.html', context)



@login_required
def patient_detail(request, pk):
    patient = get_object_or_404(Patient.objects.select_related('attending_physician', 'created_by'), pk=pk)
    # Проверка прав на просмотр этого пациента
    if not patient.user_can_view(request.user):
        messages.error(request, 'У вас нет прав для просмотра этой карты')
        return redirect('patients:patient_list')
    # Получаем историю госпитализаций
    hospitalizations = patient.hospitalizations.all()
    # Определяем доступные действия
    can_edit = patient.user_can_edit(request.user)
    can_delete = patient.user_can_delete(request.user)
    can_discharge = (patient.status == 'HOSPITALIZED' and can_edit)
    context = {
        'patient': patient,
        'hospitalizations': hospitalizations,
        'can_edit': can_edit,
        'can_delete': can_delete,
        'can_discharge': can_discharge,
    }
    return render(request, 'patients/patient_detail.html', context)


from django.views.decorators.http import require_http_methods

@login_required
@require_http_methods(["GET", "POST"])
def patient_discharge(request, pk):
    """Быстрая выписка пациента"""
    patient = get_object_or_404(Patient, pk=pk)
    # Проверка прав на выписку через user_can_edit
    if not (patient.status == 'HOSPITALIZED' and patient.user_can_edit(request.user)):
        messages.error(request, 'У вас нет прав для выписки этого пациента')
        return redirect('patients:patient_detail', pk=pk)
    if patient.status == 'HOSPITALIZED':
        patient.status = 'DISCHARGED'
        patient.discharge_date = timezone.now()
        patient.save()
        messages.success(request, f'Пациент {patient.full_name} выписан')
    else:
        messages.warning(request, f'Пациент уже имеет статус: {patient.get_status_display()}')
    return redirect('patients:patient_detail', pk=pk)


@login_required
def patient_export(request):
    """Экспорт пациентов"""
    # Определяем доступный queryset пациентов
    if request.user.is_administrator or request.user.has_perm('patients.view_all_patients'):
        patients_queryset = Patient.objects.all()
    elif request.user.is_doctor:
        patients_queryset = Patient.objects.filter(attending_physician=request.user)
    else:
        patients_queryset = Patient.objects.all()

    if request.method == 'POST':
        form = PatientExportForm(request.POST)
        form.fields['patients'].queryset = patients_queryset
        if form.is_valid():
            patients = form.cleaned_data['patients']
            export_format = form.cleaned_data['export_format']
            include_fields = form.cleaned_data.get('include_fields', [])
            # Проверка: пользователь может экспортировать только тех пациентов, которых может просматривать
            for patient in patients:
                if not patient.user_can_view(request.user):
                    messages.error(request, 'Вы можете экспортировать только доступных вам пациентов')
                    return redirect('patients:patient_export')
            # Подготовка данных для экспорта
            data = prepare_export_data(patients, include_fields)
            if export_format == 'csv':
                return export_csv(data, patients)
            elif export_format == 'xlsx':
                return export_excel(data, patients)
            elif export_format == 'json':
                return export_json(data, patients)
    else:
        # По умолчанию выбираем всех доступных пациентов
        initial = {
            'patients': patients_queryset,
            'include_fields': ['basic', 'documents', 'hospitalization'],
        }
        form = PatientExportForm(initial=initial)
        form.fields['patients'].queryset = patients_queryset
    context = {
        'form': form,
        'total_patients': patients_queryset.count(),
    }
    return render(request, 'patients/patient_export.html', context)


def prepare_export_data(patients, include_fields):
    """Подготовка данных для экспорта"""
    data = []
    
    for patient in patients:
        row = {}
        
        if 'basic' in include_fields:
            row.update({
                'Номер истории болезни': patient.case_number,
                'Фамилия': patient.last_name,
                'Имя': patient.first_name,
                'Отчество': patient.middle_name,
                'Пол': patient.get_gender_display(),
                'Дата рождения': patient.birth_date.strftime('%d.%m.%Y') if patient.birth_date else '',
                'Возраст': patient.age,
                'Место рождения': patient.birth_place,
                'Гражданство': patient.citizenship,
                'Адрес': patient.address,
                'Телефон': patient.phone,
            })
        
        if 'documents' in include_fields:
            row.update({
                'Серия паспорта': patient.passport_series,
                'Номер паспорта': patient.passport_number,
                'Кем выдан': patient.passport_issued_by,
                'Дата выдачи': patient.passport_issue_date.strftime('%d.%m.%Y') if patient.passport_issue_date else '',
                'ИНН': patient.inn,
                'Страховой полис': patient.insurance_policy,
            })
        
        if 'hospitalization' in include_fields:
            row.update({
                'Дата поступления': patient.admission_date.strftime('%d.%m.%Y %H:%M') if patient.admission_date else '',
                'Откуда поступил': patient.admission_from,
                'Кем доставлен': patient.delivered_by,
                'Диагноз направившего учреждения': patient.referral_diagnosis,
                'Диагноз при поступлении': patient.admission_diagnosis,
                'Код МКБ при поступлении': patient.admission_mkb_code,
                'Лечащий врач': patient.attending_physician.get_full_name() if patient.attending_physician else '',
            })
        
        if 'discharge' in include_fields:
            row.update({
                'Дата выписки': patient.discharge_date.strftime('%d.%m.%Y %H:%M') if patient.discharge_date else '',
                'Диагноз при выписке': patient.discharge_diagnosis,
                'Код МКБ при выписке': patient.discharge_mkb_code,
                'Исход заболевания': patient.get_outcome_display() if patient.outcome else '',
                'Трудоспособность': patient.get_work_capacity_display() if patient.work_capacity else '',
                'Статус': patient.get_status_display(),
            })
        
        if 'notes' in include_fields:
            row.update({
                'Примечания': patient.notes,
                'Дата создания': patient.created_at.strftime('%d.%m.%Y %H:%M') if patient.created_at else '',
                'Создал': patient.created_by.get_full_name() if patient.created_by else '',
            })
        
        data.append(row)
    
    return data


def export_csv(data, patients):
    """Экспорт в CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="patients_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    if data:
        writer = csv.DictWriter(response, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
    
    return response


def export_excel(data, patients):
    """Экспорт в Excel"""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="patients_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Пациенты"
    
    if data:
        # Заголовки
        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
            ws.column_dimensions[col_letter].width = 20
        
        # Данные
        for row_num, row_data in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row_num}'] = row_data.get(header, '')
    
    wb.save(response)
    return response


def export_json(data, patients):
    """Экспорт в JSON"""
    response = HttpResponse(
        json.dumps(data, ensure_ascii=False, indent=2),
        content_type='application/json; charset=utf-8'
    )
    response['Content-Disposition'] = f'attachment; filename="patients_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.json"'
    return response


@login_required
def api_diagnoses(request):
    """API для автодополнения диагнозов"""
    query = request.GET.get('q', '')
    
    if query:
        diagnoses = Diagnosis.objects.filter(
            Q(code__icontains=query) | Q(name__icontains=query)
        ).order_by('code')[:10]
    else:
        diagnoses = Diagnosis.objects.none()
    
    results = [
        {
            'id': d.code,
            'text': f'{d.code} - {d.name}',
            'code': d.code,
            'name': d.name,
            'description': d.description[:100] if d.description else '',
        }
        for d in diagnoses
    ]
    
    return JsonResponse({'results': results})


@login_required
def dashboard(request):
    """Дашборд с общей статистикой"""
    # Основная статистика
    total_patients = Patient.objects.count()
    hospitalized = Patient.objects.filter(status='HOSPITALIZED').count()
    discharged = Patient.objects.filter(status='DISCHARGED').count()
    
    # Статистика за последние 30 дней
    thirty_days_ago = timezone.now() - timezone.timedelta(days=30)
    recent_admissions = Patient.objects.filter(
        admission_date__gte=thirty_days_ago
    ).count()
    
    recent_discharges = Patient.objects.filter(
        discharge_date__gte=thirty_days_ago
    ).count()
    
    # Распределение по полу
    gender_stats = Patient.objects.values('gender').annotate(
        count=models.Count('id')
    )
    
    # Распределение по возрасту
    age_groups = {
        'До 18 лет': Patient.objects.filter(birth_date__gte=timezone.now() - timezone.timedelta(days=18*365)).count(),
        '18-30 лет': Patient.objects.filter(
            birth_date__lt=timezone.now() - timezone.timedelta(days=18*365),
            birth_date__gte=timezone.now() - timezone.timedelta(days=30*365)
        ).count(),
        '31-50 лет': Patient.objects.filter(
            birth_date__lt=timezone.now() - timezone.timedelta(days=30*365),
            birth_date__gte=timezone.now() - timezone.timedelta(days=50*365)
        ).count(),
        '51-70 лет': Patient.objects.filter(
            birth_date__lt=timezone.now() - timezone.timedelta(days=50*365),
            birth_date__gte=timezone.now() - timezone.timedelta(days=70*365)
        ).count(),
        'Старше 70 лет': Patient.objects.filter(
            birth_date__lt=timezone.now() - timezone.timedelta(days=70*365)
        ).count(),
    }
    
    # Последние поступления
    recent_patients = Patient.objects.select_related('attending_physician').order_by('-admission_date')[:10]
    
    can_export = (
        request.user.is_administrator or
        request.user.is_doctor or
        request.user.is_analyst or
        request.user.has_perm('patients.export_patients')
    )
    can_create = (
        request.user.is_administrator or
        request.user.is_doctor or
        request.user.is_nurse or
        request.user.is_registrar
    )
    context = {
        'total_patients': total_patients,
        'hospitalized': hospitalized,
        'discharged': discharged,
        'recent_admissions': recent_admissions,
        'recent_discharges': recent_discharges,
        'gender_stats': gender_stats,
        'age_groups': age_groups,
        'recent_patients': recent_patients,
        'can_export': can_export,
        'can_create': can_create,
    }
    return render(request, 'patients/dashboard.html', context)

class PatientListView(RoleRequiredMixin, ListView):
    """Список пациентов (классовое представление)"""
    model = Patient
    template_name = 'patients/patient_list.html'
    context_object_name = 'patients'
    paginate_by = 25
    allowed_roles = ['ADMIN', 'DOCTOR', 'NURSE', 'REGISTRAR', 'ANALYST']
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('attending_physician')
        # Фильтрация по правам доступа
        if not self.request.user.is_administrator:
            if self.request.user.is_doctor:
                queryset = queryset.filter(attending_physician=self.request.user)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Добавляем форму поиска и другие данные при необходимости
        return context


class PatientDetailView(ObjectPermissionMixin, DetailView):
    """Просмотр карты пациента (классовое представление)"""
    model = Patient
    template_name = 'patients/patient_detail.html'
    context_object_name = 'patient'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        patient = self.get_object()
        context['hospitalizations'] = patient.hospitalizations.all()
        context['can_edit'] = patient.user_can_edit(self.request.user)
        context['can_delete'] = patient.user_can_delete(self.request.user)
        context['can_discharge'] = (patient.status == 'HOSPITALIZED' and patient.user_can_edit(self.request.user))
        return context


class PatientCreateView(RoleRequiredMixin, CreateView):
    model = Patient
    form_class = PatientForm
    template_name = 'patients/patient_form.html'
    allowed_roles = ['ADMIN', 'DOCTOR', 'NURSE', 'REGISTRAR']
    success_url = reverse_lazy('patients:patient_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'Пациент успешно добавлен')
        return super().form_valid(form)


class PatientUpdateView(ObjectPermissionMixin, UpdateView):
    model = Patient
    form_class = PatientForm
    template_name = 'patients/patient_form.html'
    success_url = reverse_lazy('patients:patient_list')

    def form_valid(self, form):
        messages.success(self.request, 'Данные пациента обновлены')
        return super().form_valid(form)


class PatientDeleteView(ObjectPermissionMixin, DeleteView):
    model = Patient
    template_name = 'patients/patient_confirm_delete.html'
    success_url = reverse_lazy('patients:patient_list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Пациент удалён')
        return super().delete(request, *args, **kwargs)

# Для будущего: можно добавить CBV для discharge, export и др.