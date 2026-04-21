from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, FormView, TemplateView, View
from django.views.generic.edit import FormMixin
from django.urls import reverse_lazy
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.db.models import Q, Count
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.utils import timezone
from django.db import models
import json
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import Font

from users.mixins import RoleRequiredMixin, ObjectPermissionMixin
from .models import Patient, Diagnosis, Hospitalization
from .forms import PatientForm, PatientSearchForm, PatientExportForm


class DashboardView(RoleRequiredMixin, TemplateView):
    """Дашборд с общей статистикой (классовое представление)"""
    template_name = 'patients/dashboard.html'
    allowed_roles = ['ADMIN', 'DOCTOR', 'NURSE', 'REGISTRAR', 'ANALYST']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Основная статистика
        context['total_patients'] = Patient.objects.count()
        context['hospitalized'] = Patient.objects.filter(status='HOSPITALIZED').count()
        context['discharged'] = Patient.objects.filter(status='DISCHARGED').count()
        
        # Статистика за последние 30 дней
        thirty_days_ago = timezone.now() - timezone.timedelta(days=30)
        context['recent_admissions'] = Patient.objects.filter(
            admission_date__gte=thirty_days_ago
        ).count()
        
        context['recent_discharges'] = Patient.objects.filter(
            discharge_date__gte=thirty_days_ago
        ).count()
        
        # Распределение по полу
        context['gender_stats'] = Patient.objects.values('gender').annotate(
            count=Count('id')
        )
        
        # Распределение по возрасту
        today = timezone.now().date()
        context['age_groups'] = {
            'До 18 лет': Patient.objects.filter(
                birth_date__gte=today.replace(year=today.year - 18)
            ).count(),
            '18-30 лет': Patient.objects.filter(
                birth_date__lt=today.replace(year=today.year - 18),
                birth_date__gte=today.replace(year=today.year - 30)
            ).count(),
            '31-50 лет': Patient.objects.filter(
                birth_date__lt=today.replace(year=today.year - 30),
                birth_date__gte=today.replace(year=today.year - 50)
            ).count(),
            '51-70 лет': Patient.objects.filter(
                birth_date__lt=today.replace(year=today.year - 50),
                birth_date__gte=today.replace(year=today.year - 70)
            ).count(),
            'Старше 70 лет': Patient.objects.filter(
                birth_date__lt=today.replace(year=today.year - 70)
            ).count(),
        }
        
        # Последние поступления
        context['recent_patients'] = Patient.objects.select_related(
            'attending_physician'
        ).order_by('-admission_date')[:10]
        
        # Статистика по врачам (только для администраторов и аналитиков)
        if self.request.user.is_administrator or self.request.user.is_analyst:
            context['doctor_stats'] = Patient.objects.filter(
                attending_physician__isnull=False
            ).values(
                'attending_physician__last_name',
                'attending_physician__first_name'
            ).annotate(
                patient_count=Count('id'),
                hospitalized_count=Count('id', filter=Q(status='HOSPITALIZED'))
            ).order_by('-patient_count')[:5]
        
        # Права доступа
        context['can_export'] = (
            self.request.user.is_administrator or
            self.request.user.is_analyst or
            self.request.user.has_perm('patients.export_patients')
        )
        context['can_create'] = (
            self.request.user.is_administrator or
            self.request.user.is_doctor or
            self.request.user.is_nurse or
            self.request.user.is_registrar
        )
        
        return context


class PatientListView(RoleRequiredMixin, ListView):
    """Список пациентов с поиском и фильтрацией (классовое представление)"""
    model = Patient
    template_name = 'patients/patient_list.html'
    context_object_name = 'patients_with_perms'
    paginate_by = 25
    allowed_roles = ['ADMIN', 'DOCTOR', 'NURSE', 'REGISTRAR', 'ANALYST']
    
    def get_queryset(self):
        # Базовый queryset с учетом прав доступа
        if self.request.user.is_administrator or self.request.user.has_perm('patients.view_all_patients'):
            queryset = Patient.objects.all()
        elif self.request.user.is_doctor:
            queryset = Patient.objects.filter(attending_physician=self.request.user)
        else:
            queryset = Patient.objects.all()
        
        queryset = queryset.select_related('attending_physician')
        
        # Применяем фильтры из формы поиска
        self.search_form = PatientSearchForm(self.request.GET or None)
        
        # Сохраняем фильтры в сессии
        filters = {}
        if self.search_form.is_valid():
            query = self.search_form.cleaned_data.get('query')
            status = self.search_form.cleaned_data.get('status')
            gender = self.search_form.cleaned_data.get('gender')
            date_from = self.search_form.cleaned_data.get('date_from')
            date_to = self.search_form.cleaned_data.get('date_to')
            
            if query:
                filters['query'] = query
                queryset = queryset.filter(
                    Q(last_name__icontains=query) |
                    Q(first_name__icontains=query) |
                    Q(middle_name__icontains=query) |
                    Q(case_number__icontains=query) |
                    Q(passport_series__icontains=query) |
                    Q(passport_number__icontains=query) |
                    Q(inn__icontains=query) |
                    Q(phone__icontains=query) |
                    Q(address__icontains=query)
                )
            if status:
                filters['status'] = status
                queryset = queryset.filter(status=status)
            if gender:
                filters['gender'] = gender
                queryset = queryset.filter(gender=gender)
            if date_from:
                filters['date_from'] = date_from.strftime('%Y-%m-%d') if date_from else ''
                queryset = queryset.filter(admission_date__date__gte=date_from)
            if date_to:
                filters['date_to'] = date_to.strftime('%Y-%m-%d') if date_to else ''
                queryset = queryset.filter(admission_date__date__lte=date_to)
            
            # Сохраняем фильтры в сессии для экспорта
            self.request.session['patient_filters'] = filters
        
        # Сортировка
        sort_by = self.request.GET.get('sort', '-admission_date')
        if sort_by in ['last_name', 'admission_date', 'birth_date', 'case_number']:
            queryset = queryset.order_by(sort_by)
        elif sort_by == '-last_name':
            queryset = queryset.order_by('-last_name')
        else:
            queryset = queryset.order_by('-admission_date')
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Статистика по статусам
        status_counts = {}
        for status_code, status_name in Patient.STATUS_CHOICES:
            if self.request.user.is_administrator or self.request.user.has_perm('patients.view_all_patients'):
                status_counts[status_code] = Patient.objects.filter(status=status_code).count()
            elif self.request.user.is_doctor:
                status_counts[status_code] = Patient.objects.filter(
                    status=status_code,
                    attending_physician=self.request.user
                ).count()
            else:
                status_counts[status_code] = Patient.objects.filter(status=status_code).count()
        
        # Для каждого пациента вычисляем права
        patient_permissions = []
        for patient in context['object_list']:
            can_edit = patient.user_can_edit(self.request.user)
            can_delete = patient.user_can_delete(self.request.user)
            can_discharge = (patient.status == 'HOSPITALIZED' and can_edit)
            patient_permissions.append({
                'patient': patient,
                'can_edit': can_edit,
                'can_delete': can_delete,
                'can_discharge': can_discharge,
            })
        
        # Получаем фильтры из GET параметров для отображения
        get_params = self.request.GET
        filter_info = []
        if get_params.get('query'):
            filter_info.append(f"Поиск: {get_params['query']}")
        if get_params.get('status'):
            status_display = dict(Patient.STATUS_CHOICES).get(get_params['status'], get_params['status'])
            filter_info.append(f"Статус: {status_display}")
        if get_params.get('gender'):
            gender_display = dict(Patient.Gender.choices).get(get_params['gender'], get_params['gender'])
            filter_info.append(f"Пол: {gender_display}")
        if get_params.get('date_from'):
            filter_info.append(f"С даты: {get_params['date_from']}")
        if get_params.get('date_to'):
            filter_info.append(f"По дату: {get_params['date_to']}")
        
        context.update({
            'patients_with_perms': patient_permissions,
            'form': self.search_form if hasattr(self, 'search_form') else PatientSearchForm(),
            'total_patients': self.get_queryset().count(),
            'status_counts': status_counts,
            'sort_by': self.request.GET.get('sort', '-admission_date'),
            'can_create_patient': (
                self.request.user.is_administrator or 
                self.request.user.is_doctor or 
                self.request.user.is_nurse or 
                self.request.user.is_registrar
            ),
            'active_filters': bool(get_params),
            'filter_info': filter_info,
        })
        return context


class PatientDetailView(ObjectPermissionMixin, DetailView):
    """Просмотр карты пациента (классовое представление)"""
    model = Patient
    template_name = 'patients/patient_detail.html'
    context_object_name = 'patient'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.select_related('attending_physician', 'created_by')
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        patient = self.get_object()
        
        # Получаем историю госпитализаций
        hospitalizations = patient.hospitalizations.all()
        
        # Определяем доступные действия
        can_edit = patient.user_can_edit(self.request.user)
        can_delete = patient.user_can_delete(self.request.user)
        can_discharge = (patient.status == 'HOSPITALIZED' and can_edit)
        
        context.update({
            'hospitalizations': hospitalizations,
            'can_edit': can_edit,
            'can_delete': can_delete,
            'can_discharge': can_discharge,
        })
        return context
    
    def dispatch(self, request, *args, **kwargs):
        # Проверка прав на просмотр
        patient = self.get_object()
        if not patient.user_can_view(request.user):
            messages.error(request, 'У вас нет прав для просмотра этой карты')
            return redirect('patients:patient_list')
        return super().dispatch(request, *args, **kwargs)


class PatientCreateView(RoleRequiredMixin, CreateView):
    """Создание нового пациента (классовое представление)"""
    model = Patient
    form_class = PatientForm
    template_name = 'patients/patient_form.html'
    allowed_roles = ['ADMIN', 'DOCTOR', 'NURSE', 'REGISTRAR']
    success_url = reverse_lazy('patients:patient_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Пациент успешно добавлен')
        return response
    
    def get_success_url(self):
        return reverse_lazy('patients:patient_detail', kwargs={'pk': self.object.pk})


class PatientUpdateView(ObjectPermissionMixin, UpdateView):
    """Редактирование пациента (классовое представление)"""
    model = Patient
    form_class = PatientForm
    template_name = 'patients/patient_form.html'
    context_object_name = 'patient'
    
    def get_success_url(self):
        return reverse_lazy('patients:patient_detail', kwargs={'pk': self.object.pk})
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Данные пациента обновлены')
        return response


class PatientDeleteView(ObjectPermissionMixin, DeleteView):
    """Удаление пациента (классовое представление)"""
    model = Patient
    template_name = 'patients/patient_confirm_delete.html'
    success_url = reverse_lazy('patients:patient_list')
    
    def delete(self, request, *args, **kwargs):
        response = super().delete(request, *args, **kwargs)
        messages.success(self.request, 'Пациент удалён')
        return response


class PatientDischargeView(ObjectPermissionMixin, UpdateView):
    """Быстрая выписка пациента (классовое представление)"""
    model = Patient
    template_name = 'patients/patient_discharge.html'
    fields = ['discharge_date', 'discharge_diagnosis', 'discharge_mkb_code', 'outcome', 'work_capacity']
    
    def get_success_url(self):
        return reverse_lazy('patients:patient_detail', kwargs={'pk': self.object.pk})
    
    def dispatch(self, request, *args, **kwargs):
        patient = self.get_object()
        # Проверка прав на выписку через user_can_edit
        if not (patient.status == 'HOSPITALIZED' and patient.user_can_edit(request.user)):
            messages.error(request, 'У вас нет прав для выписки этого пациента')
            return redirect('patients:patient_detail', pk=patient.pk)
        return super().dispatch(request, *args, **kwargs)
    
    def form_valid(self, form):
        patient = form.save(commit=False)
        patient.status = 'DISCHARGED'
        if not patient.discharge_date:
            patient.discharge_date = timezone.now()
        patient.save()
        messages.success(self.request, f'Пациент {patient.full_name} выписан')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['patient'] = self.get_object()
        return context


class PatientQuickDischargeView(ObjectPermissionMixin, View):
    """Быстрая выписка пациента без формы (классовое представление)"""
    
    def post(self, request, pk):
        patient = get_object_or_404(Patient, pk=pk)
        
        # Проверка прав на выписку
        if not (patient.status == 'HOSPITALIZED' and patient.user_can_edit(request.user)):
            messages.error(request, 'У вас нет прав для выписки этого пациента')
            return redirect('patients:patient_detail', pk=pk)
        
        patient.status = 'DISCHARGED'
        patient.discharge_date = timezone.now()
        patient.save()
        
        messages.success(request, f'Пациент {patient.full_name} выписан')
        return redirect('patients:patient_detail', pk=pk)


class PatientExportView(RoleRequiredMixin, FormView):
    """Экспорт пациентов (классовое представление)"""
    template_name = 'patients/patient_export.html'
    form_class = PatientExportForm
    allowed_roles = ['ADMIN', 'DOCTOR', 'NURSE', 'REGISTRAR', 'ANALYST']
    
    def get_initial(self):
        initial = super().get_initial()
        
        # Получаем queryset пациентов с учетом фильтров из сессии
        patients_queryset = self._get_filtered_patients_queryset()
        
        initial.update({
            'patients': patients_queryset,
            'include_fields': ['basic', 'documents', 'hospitalization'],
            'export_format': 'xlsx',
        })
        return initial
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        
        # Передаем queryset для поля patients
        if 'data' not in kwargs:
            kwargs['initial'] = self.get_initial()
        
        # Обновляем queryset для поля patients
        form = PatientExportForm(**kwargs)
        form.fields['patients'].queryset = self._get_filtered_patients_queryset()
        
        return kwargs
    
    def _get_filtered_patients_queryset(self):
        """Получаем queryset пациентов с учетом фильтров из сессии"""
        # Пытаемся получить фильтры из сессии
        filters = self.request.session.get('patient_filters', {})
        
        # Базовый queryset с учетом прав доступа
        if self.request.user.is_administrator or self.request.user.has_perm('patients.view_all_patients'):
            queryset = Patient.objects.all()
        elif self.request.user.is_doctor:
            queryset = Patient.objects.filter(attending_physician=self.request.user)
        else:
            queryset = Patient.objects.all()
        
        # Применяем фильтры из сессии
        if filters:
            query = filters.get('query')
            status = filters.get('status')
            gender = filters.get('gender')
            date_from = filters.get('date_from')
            date_to = filters.get('date_to')
            
            if query:
                queryset = queryset.filter(
                    Q(last_name__icontains=query) |
                    Q(first_name__icontains=query) |
                    Q(middle_name__icontains=query) |
                    Q(case_number__icontains=query) |
                    Q(passport_series__icontains=query) |
                    Q(passport_number__icontains=query) |
                    Q(inn__icontains=query) |
                    Q(phone__icontains=query) |
                    Q(address__icontains=query)
                )
            if status:
                queryset = queryset.filter(status=status)
            if gender:
                queryset = queryset.filter(gender=gender)
            if date_from:
                queryset = queryset.filter(admission_date__date__gte=date_from)
            if date_to:
                queryset = queryset.filter(admission_date__date__lte=date_to)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        patients_queryset = self._get_filtered_patients_queryset()
        context['total_patients'] = patients_queryset.count()
        
        # Добавляем информацию о фильтрах
        filters = self.request.session.get('patient_filters', {})
        context['active_filters'] = bool(filters)
        context['filter_info'] = self._get_filter_info(filters)
        
        return context
    
    def _get_filter_info(self, filters):
        """Формируем информацию о примененных фильтрах"""
        if not filters:
            return []
        
        info = []
        if filters.get('query'):
            info.append(f"Поиск: {filters['query']}")
        if filters.get('status'):
            status_display = dict(Patient.STATUS_CHOICES).get(filters['status'], filters['status'])
            info.append(f"Статус: {status_display}")
        if filters.get('gender'):
            gender_display = dict(Patient.Gender.choices).get(filters['gender'], filters['gender'])
            info.append(f"Пол: {gender_display}")
        if filters.get('date_from'):
            info.append(f"С даты: {filters['date_from']}")
        if filters.get('date_to'):
            info.append(f"По дату: {filters['date_to']}")
        
        return info
    
    def form_valid(self, form):
        patients = form.cleaned_data['patients']
        export_format = form.cleaned_data['export_format']
        include_fields = form.cleaned_data.get('include_fields', [])
        
        # Проверка: пользователь может экспортировать только тех пациентов, которых может просматривать
        for patient in patients:
            if not patient.user_can_view(self.request.user):
                messages.error(self.request, 'Вы можете экспортировать только доступных вам пациентов')
                return self.form_invalid(form)
        
        # Подготовка данных для экспорта
        data = self._prepare_export_data(patients, include_fields)
        
        if export_format == 'csv':
            return self._export_csv(data)
        elif export_format == 'xlsx':
            return self._export_excel(data)
        elif export_format == 'json':
            return self._export_json(data)
        
        return super().form_valid(form)

    def _prepare_export_data(self, patients, include_fields):
        """Подготовка данных для экспорта"""
        data = {
            'headers': [],
            'rows': [],
            'count': len(patients),
        }

        # Определяем, какие поля включать
        all_fields = []
        
        if 'basic' in include_fields:
            all_fields.extend([
                ('case_number', 'ИБ №'),
                ('last_name', 'Фамилия'),
                ('first_name', 'Имя'),
                ('middle_name', 'Отчество'),
                ('birth_date', 'Дата рождения'),
                ('age', 'Возраст'),
                ('gender', 'Пол'),
                ('phone', 'Телефон'),
            ])
        
        if 'documents' in include_fields:
            all_fields.extend([
                ('passport_series', 'Серия паспорта'),
                ('passport_number', 'Номер паспорта'),
                ('passport_issued_by', 'Кем выдан'),
                ('passport_issue_date', 'Дата выдачи'),
                ('inn', 'ИНН'),
                ('insurance_policy', 'Полис'),
                ('address', 'Адрес'),
            ])
        
        if 'hospitalization' in include_fields:
            all_fields.extend([
                ('admission_date', 'Дата поступления'),
                ('admission_from', 'Откуда поступил'),
                ('admission_diagnosis', 'Диагноз при поступлении'),
                ('admission_mkb_code', 'МКБ при поступлении'),
                ('status', 'Статус'),
                ('attending_physician', 'Лечащий врач'),
            ])
        
        if 'discharge' in include_fields:
            all_fields.extend([
                ('discharge_date', 'Дата выписки'),
                ('discharge_diagnosis', 'Диагноз при выписке'),
                ('discharge_mkb_code', 'МКБ при выписке'),
                ('outcome', 'Исход'),
            ])
        
        if 'notes' in include_fields:
            all_fields.extend([
                ('notes', 'Примечания'),
            ])
        
        # Добавляем заголовки
        data['headers'] = [field[1] for field in all_fields]
        
        # Добавляем данные пациентов
        for patient in patients:
            row = []
            for field_name, field_label in all_fields:
                value = getattr(patient, field_name, '')
                
                # Обработка специальных полей
                if field_name == 'age':
                    value = patient.age
                elif field_name == 'gender':
                    value = patient.get_gender_display()
                elif field_name == 'status':
                    value = patient.get_status_display()
                elif field_name == 'outcome':
                    value = patient.get_outcome_display() if patient.outcome else ''
                elif field_name == 'attending_physician':
                    if patient.attending_physician:
                        value = patient.attending_physician.get_full_name() or patient.attending_physician.username
                    else:
                        value = ''
                elif value is None:
                    value = ''
                
                row.append(str(value))
            
            data['rows'].append(row)
        
        return data
    
    def _export_csv(self, data):
        """Экспорт в CSV"""
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="patients_export.csv"'
        
        writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_MINIMAL)
        
        # Записываем заголовки
        writer.writerow(data['headers'])
        
        # Записываем данные
        for row in data['rows']:
            writer.writerow(row)
        
        return response

    def _export_excel(self, data):
        """Экспорт в Excel"""
        from io import BytesIO
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = "Пациенты"
        
        # Создаем шрифт для заголовков
        header_font = Font(bold=True)
        
        # Записываем заголовки
        for col_num, header in enumerate(data['headers'], 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
        
        # Записываем данные
        for row_num, row_data in enumerate(data['rows'], 2):
            for col_num, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=cell_value)
        
        # Настраиваем ширину колонок
        for col_num, header in enumerate(data['headers'], 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 20
        
        # Сохраняем в буфер
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="patients_export.xlsx"'
        
        return response
    
    def _export_json(self, data):
        """Экспорт в JSON"""
        from django.core.serializers.json import DjangoJSONEncoder
        
        export_data = {
            'headers': data['headers'],
            'data': data['rows'],
            'count': data['count'],
            'export_date': timezone.now().isoformat(),
        }
        
        response = HttpResponse(
            json.dumps(export_data, ensure_ascii=False, indent=2, cls=DjangoJSONEncoder),
            content_type='application/json; charset=utf-8'
        )
        response['Content-Disposition'] = 'attachment; filename="patients_export.json"'
        
        return response


class PatientClearFiltersView(LoginRequiredMixin, View):
    """Очистка фильтров пациентов из сессии"""
    
    def get(self, request):
        if 'patient_filters' in request.session:
            del request.session['patient_filters']
            messages.success(request, 'Фильтры очищены')
        return redirect('patients:patient_list')


class ApiDiagnosesView(LoginRequiredMixin, View):
    """API для автодополнения диагнозов (классовое представление)"""
    
    def get(self, request):
        query = request.GET.get('q', '')
        
        if query:
            diagnoses = Diagnosis.objects.filter(
                Q(code__icontains=query) | Q(name__icontains=query)
            ).order_by('code')[:10]
        else:
            diagnoses = Diagnosis.objects.none()
        
        results = [
            {
                'id': d.code,
                'text': f'{d.code} - {d.name}',
                'code': d.code,
                'name': d.name,
                'description': d.description[:100] if d.description else '',
            }
            for d in diagnoses
        ]
        
        return JsonResponse({'results': results})


class PatientStatisticsView(RoleRequiredMixin, TemplateView):
    """Расширенная статистика пациентов (классовое представление)"""
    template_name = 'patients/statistics.html'
    allowed_roles = ['ADMIN', 'DOCTOR', 'ANALYST']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Фильтры
        year = self.request.GET.get('year', timezone.now().year)
        month = self.request.GET.get('month')
        
        # Базовый queryset
        queryset = Patient.objects.filter(
            admission_date__year=year
        )
        
        if month:
            queryset = queryset.filter(admission_date__month=month)
        
        # Общая статистика
        context['total_patients'] = queryset.count()
        context['hospitalized'] = queryset.filter(status='HOSPITALIZED').count()
        context['discharged'] = queryset.filter(status='DISCHARGED').count()
        context['transferred'] = queryset.filter(status='TRANSFERRED').count()
        context['died'] = queryset.filter(status='DIED').count()
        
        # Статистика по месяцам
        monthly_stats = []
        for month_num in range(1, 13):
            month_data = queryset.filter(admission_date__month=month_num)
            monthly_stats.append({
                'month': month_num,
                'admissions': month_data.count(),
                'discharges': month_data.filter(discharge_date__month=month_num).count(),
            })
        context['monthly_stats'] = monthly_stats
        
        # Статистика по диагнозам (топ-10)
        context['top_diagnoses'] = Diagnosis.objects.annotate(
            patient_count=Count('patient')
        ).order_by('-patient_count')[:10]
        
        # Статистика по возрасту
        age_stats = {
            'До 18 лет': queryset.filter(age__lt=18).count(),
            '18-30 лет': queryset.filter(age__range=(18, 30)).count(),
            '31-50 лет': queryset.filter(age__range=(31, 50)).count(),
            '51-70 лет': queryset.filter(age__range=(51, 70)).count(),
            'Старше 70 лет': queryset.filter(age__gt=70).count(),
        }
        context['age_stats'] = age_stats
        
        # Годы для фильтра
        context['years'] = Patient.objects.dates('admission_date', 'year')
        
        return context


class HospitalizationCreateView(RoleRequiredMixin, CreateView):
    """Добавление госпитализации (классовое представление)"""
    model = Hospitalization
    template_name = 'patients/hospitalization_form.html'
    fields = ['admission_date', 'diagnosis', 'mkb_code', 'department', 'attending_physician', 'notes']
    allowed_roles = ['ADMIN', 'DOCTOR', 'NURSE']
    
    def dispatch(self, request, *args, **kwargs):
        self.patient = get_object_or_404(Patient, pk=kwargs['patient_pk'])
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['patient'] = self.patient
        return context
    
    def form_valid(self, form):
        hospitalization = form.save(commit=False)
        hospitalization.patient = self.patient
        hospitalization.save()
        messages.success(self.request, 'Госпитализация добавлена')
        return redirect('patients:patient_detail', pk=self.patient.pk)


class HospitalizationUpdateView(ObjectPermissionMixin, UpdateView):
    """Редактирование госпитализации (классовое представление)"""
    model = Hospitalization
    template_name = 'patients/hospitalization_form.html'
    fields = ['admission_date', 'discharge_date', 'diagnosis', 'mkb_code', 'department', 
              'attending_physician', 'outcome', 'notes']
    
    def get_success_url(self):
        return reverse_lazy('patients:patient_detail', kwargs={'pk': self.object.patient.pk})
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Госпитализация обновлена')
        return response


class HospitalizationDeleteView(ObjectPermissionMixin, DeleteView):
    """Удаление госпитализации (классовое представление)"""
    model = Hospitalization
    template_name = 'patients/hospitalization_confirm_delete.html'
    
    def get_success_url(self):
        return reverse_lazy('patients:patient_detail', kwargs={'pk': self.object.patient.pk})
    
    def delete(self, request, *args, **kwargs):
        response = super().delete(request, *args, **kwargs)
        messages.success(self.request, 'Госпитализация удалена')
        return response